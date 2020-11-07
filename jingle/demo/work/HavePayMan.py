# coding: utf-8
# 每月已交农保计数统计

import openpyxl,os

class village:
    def __init__(self, id,name,count = 0):
        self.id = id
        self.name = name
        self.thisYearCount = count
        self.passYearPayCount = 0
villages = [
    village('4402820401','赤岭村',0),
    village('4402820402','弱过村',0),
    village('4402820403','大部村',0),
    village('4402820404','云西村',0),
    village('4402820405','大坪村',0),
    village('4402820406','沙头村',0),
    village('4402820407','下湖村',0),
    village('4402820408','河村村',0),
    village('4402820409','下楼村',0),
    village('4402820410','石庄村',0),
    village('4402820411','群星村',0),
    village('4402820412','泷头村',0),
    village('4402820413','水口村',0),
    village('4402820414','社区',0)
]
idToVillage = {}
for v in villages:
    idToVillage[v.id] = v

# xlsName = '12月已缴农保_20191223'
# txtPath = dir + xlsName + '.txt'
dir = 'C:/Users/Administrator/Desktop/农保未续缴明细/已缴农保/'
statisticsFileName = dir + '已缴人员汇总统计表2020.xlsx'
xsl = openpyxl.load_workbook(statisticsFileName)
# xsl = openpyxl.Workbook()

# 文件路径，写入的列
def analyseOneFile(txtPath,colNum = 15):
    txtFile = open(txtPath,'rb') #因为有乱码，用正常文本解析会报错，只能用二进制解析
    for line in txtFile:
        # print(line)
        if '#' in str(line[:10], encoding="gbk"): continue
        recordArr = line.split(b'|')
        # print(str(recordArr[12], encoding="gbk"))
        theId = str(recordArr[1], encoding="gbk")
        theVillage = idToVillage[theId]
        if '期缴' == str(recordArr[12], encoding="gbk"):
            theVillage.thisYearCount = theVillage.thisYearCount + 1
        else:
            pass # 不统计补缴
            # theVillage.passYearPayCount = theVillage.passYearPayCount + 1

    for k,v in idToVillage.items():
        print(v.name, v.thisYearCount)

    # 写入数据
    sheet = xsl.worksheets[0]
    for rowNum in range(4,18): # 按villages定义的顺序写入数据
        sheet.cell(rowNum, colNum+1).value = villages[rowNum-4].thisYearCount
        sheet.cell(rowNum+16, colNum + 1).value = villages[rowNum - 4].passYearPayCount
    # 重置计数器
    for v in villages:
        v.passYearPayCount = 0
        v.thisYearCount = 0

for oneTxtFile in os.listdir(dir):# 每月一个文件，oneTxtFile的文件名例子：06月已缴农保_20191223.txt
    # print(oneTxtFile,oneTxtFile[-3:],oneTxtFile[:3])
    if oneTxtFile[-3:]== 'txt': # 如果是文本文件才解析
        # 根据月份设置写入的行
        analyseOneFile(dir + '/' + oneTxtFile,int(oneTxtFile[:2]))

xsl.save(statisticsFileName)