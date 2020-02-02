from openpyxl.styles import Border, Side
import openpyxl

class village:
    def __init__(self, id,name,count = 0,obj = None,remark=''):
        self.id = id
        self.name = name
        self.count = count
        self.obj = obj
        self.remark = remark
    def __repr__(self):
        return '{%s,%s,%s}'%(self.id,self.name,self.count)
villages = [
    village('4402820401','赤岭',0),
    village('4402820402','弱过',0),
    village('4402820403','大部',0),
    village('4402820404','云西',0),
    village('4402820405','大坪',0),
    village('4402820406','沙头',0),
    village('4402820407','下湖',0),
    village('4402820408','河村',0),
    village('4402820409','下楼',0),
    village('4402820410','石庄',0),
    village('4402820411','群星',0),
    village('4402820412','泷头',0),
    village('4402820413','水口',0),
    village('4402820414','社区',0)
]
nameToVillage = {}
for v in villages:
    nameToVillage[v.name] = v

class ColProp:
    def __init__(self, head,len = 8,obj = None,remark='',**argDict):
        self.head = head
        self.len = len
        self.obj = obj
        self.remark = remark
        for k,v in argDict.items():
            setattr(self,k,v)

# 设置边框和列宽
thin = Side(border_style="thin", color="000000")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
# 如果含有以下关键词，就设置对应列宽
colHeadToWidth = {'序号':4,'村':20,'姓名':8,'性别':4,'身份':20,'月数':8,'档次':24,'银行':20}
def setBorderWidth(sheet,maxBorderRowNum=None):
    for rowI,row in enumerate(sheet):
        for cellIdx,cell in enumerate(row):
            cell.border = border # 画边框
            if rowI == 0:# 获取列头，判断该使用什么列宽
                cp = ColProp(cell.value)
                for k,v in colHeadToWidth.items():
                    if k in cp.head:
                        cp.len = v
                        cp.remark = openpyxl.utils.get_column_letter(cellIdx+1)
                        sheet.column_dimensions[cp.remark].width = cp.len
                        # print(cp.head,cp.len)


    if maxBorderRowNum == None or sheet.max_column > maxBorderRowNum:
        return
    # 如果设置了要画边框的最大行，而且数据行数小于设置的行数，那么继续画边框
    for ri in range(1,maxBorderRowNum + 1):
        for ci in range(1,sheet.max_column + 1):
            sheet.cell(ri,ci).border = border
