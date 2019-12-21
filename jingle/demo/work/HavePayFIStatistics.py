# coding: utf-8
# 每月已交农保计数统计

import openpyxl

class village:
    def __init__(self, id,name,count = 0):
        self.id = id
        self.name = name
        self.count = count
villages = [
    village('4402820401','赤岭村',0),
    village('4402820402','弱过村',0),
    village('4402820403','大部村',0),
    village('4402820404','云西村',0),
    village('4402820405','大坪村',0),
    village('4402820406','沙头村',0),
    village('4402820407','下湖村',0),
    village('4402820408','河村村',0),
    village('4402820409','下楼村',0),
    village('4402820410','石庄村',0),
    village('4402820411','群星村',0),
    village('4402820412','泷头村',0),
    village('4402820413','水口村',0),
    village('4402820414','社区',0)
]
idToVillage = {}
for v in villages:
    idToVillage[v.id] = v


xlsName = '12月已缴农保_20191220'
dir = 'C:/Users/Administrator/Desktop/农保未续缴明细/'
statisticsFileName = 'C:/Users/Administrator/Desktop/杨欢欢/新农保年已交费人员/已缴人员汇总统计表2019（全）.xlsx'
txtPath = dir + xlsName + '.txt'
txtFile = open(txtPath,'rb') #因为有乱码，用正常文本解析会报错，只能用二进制解析


xsl = openpyxl.load_workbook(statisticsFileName)
for line in txtFile:
    if '#' in str(line[:10], encoding="gbk"): continue
    # print(str(line[:10], encoding="gbk"))
    recordArr = line.split(b'\t')
    theId = str(recordArr[1], encoding="gbk")
    theVillage = idToVillage[theId]
    theVillage.count = theVillage.count +1
for k,v in idToVillage.items():
    print(v.name,v.count)
sheet = xsl.worksheets[0]
for rowNum in range(4,17):
    print(rowNum,sheet.cell(rowNum,13).value)
    sheet.cell(rowNum, 15).value = villages[rowNum-4].count
xsl.save(statisticsFileName)