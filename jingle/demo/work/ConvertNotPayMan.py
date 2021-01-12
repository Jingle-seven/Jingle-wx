# coding: utf-8
# 未续缴农保名单，txt转excel，删掉0

import openpyxl,os,time
from openpyxl.styles import Border, Side
import jingle.demo.work.SkData as SkData

dir = 'C:/Users/Administrator/Desktop/农保未续缴明细/'
strTime = time.strftime("%Y%m%d", time.localtime())

wbook = openpyxl.Workbook()
wbook.remove(wbook['Sheet'])# 删除默认的Sheet
idToAddress  = {} #地址信息
def readIdToAddressInfo():
    kv = dict()
    villageBook = openpyxl.load_workbook('C:/Users/Administrator/Desktop/常用文件/部门数据.xlsx')
    for row in villageBook.worksheets[0].values:
        address = row[0]
        if len(address.split('村委会'))<2:
            continue
        kv[row[6]] = address.split('村委会')[1]
    return kv

def convertOneVillage(txtFullPath,destFileName=None):
    # txtPath = dir + countryName +'.txt'
    txtFile = open(txtFullPath, encoding='gbk')
    filePath, txtName = os.path.split(txtFullPath) # 分离路径和文件名
    countryName = str(txtName.split('.')[0]) # 分离村名
    sheet = wbook.create_sheet(countryName[:2],0)
    tableHead = ['序号', '姓名', '身份证('+countryName[:2]+')', '档次', '账号', '已交月份']
    sheet.append(tableHead)
    # 数据
    rowNum = 2
    for line in txtFile:
        what = line.split('|')
        # print(type(what),what)
        if line.startswith('#'): continue
        if what[13] == '0': continue #删掉已交0期的数据
        sheet.cell(rowNum, 1).value = rowNum - 1
        sheet.cell(rowNum, 2).value = what[2]
        sheet.cell(rowNum, 3).value = what[3]
        sheet.cell(rowNum, 4).value = what[6]
        sheet.cell(rowNum, 5).value = what[10]
        sheet.cell(rowNum, 6).value = what[13]
        rowNum  = rowNum + 1
    # 边框
    thin = Side(border_style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in sheet['A1:F%s'%(rowNum-1)]:
        for cell in row:
            cell.border = border
    # 列宽
    sheet.column_dimensions['A'].width = 4
    sheet.column_dimensions['B'].width = 8
    sheet.column_dimensions['C'].width = 20
    sheet.column_dimensions['D'].width = 22
    sheet.column_dimensions['E'].width = 20
    sheet.column_dimensions['F'].width = 8

    print('%s %s end'%(countryName,rowNum-2))
    destFullFilePath = dir + countryName +'.xlsx'
    if destFileName != None: # 如果规定了文件名，就不用txt文件名作为文件名，用于多个村一个表
        destFullFilePath = dir + destFileName + '.xlsx'
    wbook.save(destFullFilePath)
    print(destFullFilePath)

def convertManyVillage(txtFullPath):
    destDir = txtFullPath.split('.')[0]
    if not os.path.exists(destDir):
        os.mkdir(destDir)
    txtFile = open(txtFullPath, encoding='gbk')
    tableHead = ['序号', '村委会','姓名', '身份证号', '未交农保档次', '银行账号', '已交月数','村小组']
    for v in SkData.villages:
        v.book = openpyxl.Workbook() # 每个村单独的excel文件
        v.book.worksheets[0].append(tableHead)
        v.sumSheet = wbook.create_sheet(v.name) # 一个excel文件，每个村一个工作表
        v.sumSheet.append(tableHead)

    # 遍历每一行
    for line in txtFile:
        r = line.split('|')
        if line.startswith('#'): continue
        if r[13] == '0': continue #删掉已交0期的数据，有可能是特殊参保人群
        # print(type(r), r[8])
        v = SkData.idToVillage[r[8]]
        v.count = v.count +1
        bankAccount = '未绑定'
        address = idToAddress.get(r[3],'未找到')
        stage = r[6].replace('按年缴费第','').replace(':','')
        if r[10]:
            bankAccount = '***' + r[10][-5:-1]
        v.book.worksheets[0].append([v.count,v.name,r[2],r[3],stage,r[10],r[13],address]) # 单独村文件不打码银行账号，方便村干部代存钱
        v.sumSheet.append([v.count,v.name,r[2],r[3][0:10]+'***',stage,bankAccount,r[13],address])

    countAll = 0
    for v in SkData.villages:
        SkData.setBorderWidth(v.book.worksheets[0], 25) # 设置边框直到25行
        SkData.setBorderWidth(v.sumSheet, 25)  # 设置边框直到25行
        v.book.save(destDir +'/'+v.name+'本年未交农保'+strTime+'.xlsx')
        countAll = countAll + v.count
    wbook.save(destDir +'/水口镇各村本年未续交农保'+strTime+'.xlsx')
    print('转换了%s条数据'%countAll)

if __name__ == '__main__':
    print('启动')
    idToAddress = readIdToAddressInfo() # 读取地址信息
    print('读取地址完毕')
    fileName = '水口镇未续缴农保_20201225.txt'
    fullFilePath = dir + fileName
    if os.path.isdir(fullFilePath): # 如果是目录，遍历目录下的所有文件转成excel
        for oneTxtFile in os.listdir(fullFilePath):
            convertOneVillage(fullFilePath + '/' + oneTxtFile, fileName)
    else:
        # convertOneVillage(fullFilePath) # 只转换一个村，一个文本文件
        convertManyVillage(fullFilePath)  # 转换多村，一个文本文件
    print('完成')