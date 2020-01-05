# coding: utf-8
# 未续缴农保名单，txt转excel，删掉0

import openpyxl,os
from openpyxl.styles import Border, Side

dir = 'C:/Users/Administrator/Desktop/农保未续缴明细/'
wbook = openpyxl.Workbook()
# sheet = wbook.create_sheet('s1')
wbook.remove(wbook['Sheet'])# 删除默认的Sheet
def convertOneVillage(txtFullPath,destFileName=None):
    # txtPath = dir + countryName +'.txt'
    txtFile = open(txtFullPath, encoding='gbk')
    filePath, txtName = os.path.split(txtFullPath) # 分离路径和文件名
    countryName = str(txtName.split('.')[0]) # 分离村名
    sheet = wbook.create_sheet(countryName[:2],0)
    tableHead = ['序号', '姓名', '身份证('+countryName[:2]+')', '档次', '账号', '已交月份']
    sheet.append(tableHead)
    # 数据
    rowNum = 2
    for line in txtFile:
        what = line.split('|')
        # print(type(what),what)
        if line.startswith('#'): continue
        if what[13] == '0': continue #删掉已交0期的数据
        sheet.cell(rowNum, 1).value = rowNum - 1
        sheet.cell(rowNum, 2).value = what[2]
        sheet.cell(rowNum, 3).value = what[3]
        sheet.cell(rowNum, 4).value = what[6]
        sheet.cell(rowNum, 5).value = what[10]
        sheet.cell(rowNum, 6).value = what[13]
        rowNum  = rowNum + 1
    # 边框
    thin = Side(border_style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in sheet['A1:F%s'%(rowNum-1)]:
        for cell in row:
            cell.border = border
    # 列宽
    sheet.column_dimensions['A'].width = 4
    sheet.column_dimensions['B'].width = 8
    sheet.column_dimensions['C'].width = 20
    sheet.column_dimensions['D'].width = 22
    sheet.column_dimensions['E'].width = 20
    sheet.column_dimensions['F'].width = 8

    print('%s %s end'%(countryName,rowNum-2))
    destFullFilePath = dir + countryName +'.xlsx'
    if destFileName != None: # 如果规定了文件名，就不用txt文件名作为文件名，用于多个村一个表
        destFullFilePath = dir + destFileName + '.xlsx'
    wbook.save(destFullFilePath)
    print(destFullFilePath)

if __name__ == '__main__':
    # fileName = '赤岭未缴农保_20191226.txt'
    fileName = '水口镇未续缴农保_20191227'
    fullFilePath = dir + fileName
    if os.path.isdir(fullFilePath): # 如果是目录，遍历目录下的所有文件转成excel
        for oneTxtFile in os.listdir(fullFilePath):
            convertOneVillage(fullFilePath + '/' + oneTxtFile, fileName)
    else:
        convertOneVillage(fullFilePath)
    print('ok')