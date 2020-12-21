# coding: utf-8
# 把各种excel分村委，按村委分成多个sheet

import openpyxl,time
import jingle.demo.work.SkData as Data
from openpyxl.styles import Border, Side
# from openpyxl.utils import get_column_letter


startTs = time.time()
# dir = 'C:/Users/Administrator/Desktop/农保未续缴明细/'
dir='C:/Users/Administrator/Desktop/'
xlsName = '水口镇上年已参本年未参医保名单1206'
skColHead = ''
skColValue = '水口镇'
separateColHead = '社区村'
skColHeadIdx = 0
separateColHeadIdx = 0
# 此处设置列宽已经无效的，要在文件末尾单独设置
colHeadToWidth = {'序号':4, '公安户籍编号':9, '姓名':10, '公民身份号码':6, '户籍地址':20, '社区村':20,'本年度不参保原因':24}
# colHeadToWidth = {'序号':4,'村委':20,'姓名':8,'性别':8,'身份证号':20,'居住地址':0,
#                   '人社局下发火化日期（参考）':14,'人社所历年上报死亡日期（参考）':16,'死亡日期':16}
nameToColProp = {}
for k,v in colHeadToWidth.items():
    nameToColProp[k] = Data.ColProp(k,v,remark=-1)

rawBook = openpyxl.load_workbook(dir + xlsName + '.xlsx')
resBook = openpyxl.Workbook()
resBook.remove(resBook.worksheets[0])
nameToVillage = {}
for v in Data.villages: #创建工作表，并保存工作表的引用。写备注,创字典，便于后面匹配
    v.obj = resBook.create_sheet(v.name)
    v.obj.append(list(colHeadToWidth.keys())) #表头
    if v.name == '社区':
        v.remark = '水口社区居委会'#这个标记根据待处理文件中的居委具体名称而定
    else:
        v.remark = v.name +'村委会'#这个标记根据待处理文件中村委的具体名称而定
        # v.remark = v.name +'村委会'
    nameToVillage[v.remark] = v
print(rawBook.sheetnames)
nowSheet = rawBook.worksheets[0]
counter = 0
# for i in range(1,nowSheet.max_row):
#     row = nowSheet[i] # 用下标取值效率过低，耗时是values方法的十倍，可能是因为要构造大量cell对象
    # rowValue = [cell.value for cell in s1[i]] # s2[i]是一行元素为Cell的元组，转换为value的list更直接，便于使用
for row in nowSheet.values:
    counter += 1
    if counter==1:
        if len(skColHead) > 0:# 如果有水口镇过滤条件,找出水口镇过滤条件的那一列的序号，存入skTbHeadIdx
            skColHeadIdx = row.index(skColHead)
        separateColHeadIdx = row.index(separateColHead)# .分类条件，存入separateColHeadIdx
        for k,v in nameToColProp.items():# 找出所需列的下标
            if k == '序号': # 原有的序号列，不加到所需列表
                continue
            try:
                v.remark = row.index(k)
            except Exception as e:
                print('表头没有 {} 列'.format(k),e)
                pass
    # print(row[skColHeadIdx],row[skColHeadIdx].find(skColValue))
    # 如果没有设置水口镇过滤器
    # 或者设置了，并且是水口镇，将数据写入到对应村委工作表
    if len(skColHead)==0 or row[skColHeadIdx].find(skColValue) >= 0:
        try:#
            # print(row[separateColHeadIdx])
            village = nameToVillage[row[separateColHeadIdx]]
        except Exception as e:
            continue
        village.count +=1
        resRow = [village.count]
        # print(village,village.count)
        for k,v in nameToColProp.items():
            if v.remark==-1: # 如果是原表格中没有的列
                if k=='序号': # 而且不是第一列序号列，那么填空白
                    pass
                else:# 如果是原表格中没有的列,而且是序号列，跳过
                    resRow.append('')
            else:
                resRow.append(row[v.remark])
        village.obj.append(resRow)
    # print(row)

# 设置边框和列宽
cToW = {'身份证号':14,'户籍地':22,'本年度不参保原因':12}
for k,v in nameToVillage.items():
    Data.setBorderWidth(v.obj,specifiedColWidth=cToW)


print('处理了{}行数据，耗时{:.2f}秒'.format(counter,time.time() - startTs))
print(dir + xlsName + '-水口各村.xlsx')
resBook.save(dir + xlsName + '-水口各村.xlsx')