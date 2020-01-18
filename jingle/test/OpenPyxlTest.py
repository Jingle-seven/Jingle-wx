# coding: utf-8

import openpyxl

wbook = openpyxl.Workbook()
sheet = wbook.create_sheet('sheet1',0)# 新建工作表（表名，位置）
for sheet in wbook: #遍历工作表
    print(sheet.title)
print(wbook.sheetnames) #工作表名
sheet = wbook['Sheet'] # 按名字获取工作表
sheet2 = wbook.worksheets[0] # 按顺序获取工作表
sheet.title = 'new_sheet' # 更改表名
wbook.remove(wbook.worksheets[0]) #删除工作表

print(sheet.max_row,sheet.max_column) # 工作表最大行列
tableHead = ['序号', '姓名', '身份证', '档次', '账号', '已交月份']
for i in tableHead:
    sheet.cell(1, tableHead.index(i) + 1).value = i # 表头
sheet.append([]) #插入空白行
sheet.append(tableHead) #插入数据行
data = [[1,2,3],[4,5,6]] # 按行填充数据
for x in data:
    sheet.append(x)

sheet['A4'] = 4 #给第4行第A列的单元格赋值为4
sheet.cell(row=4, column=2, value=10) #给第4行第2列的单元格赋值为10
sheet.cell(4, 2, 10) #同上
cell = sheet.cell(4,2)
cell.value = 'hello, world'
print(sheet.cell(4,2).value)


# 遍历值
for row in sheet.values:
   for value in row:
     print(value)
for row in sheet.iter_rows(min_row=1, max_col=3, max_row=2):
        for cell in row:
            print(cell)

# 获取单元格类型，如果是常规，显示general,如果是数字，显示'0.00_ '，如果是百分数显示0%
# 数字需要在Excel中设置数字类型，直接写入的数字是常规类型
print(sheet["A4"].number_format)
#公式，打印的是公式内容，不是公式计算后的值,程序无法取到计算后的值
sheet["A5"] = "=SUM(A1:A3)"
print(sheet["A4"].value)
#合并后的单元格，脚本单独执行拆分操作会报错，需要重新执行合并操作再拆分
sheet.merge_cells('A2:D2')
sheet.unmerge_cells('A2:D2')
#在第7行之上插入一行
sheet.insert_rows(7)
#从第6列开始，删除3列，即删除6、7、8列，如下：
sheet.delete_cols(6, 3)

# 字体样式
from openpyxl.styles import Font
font = Font(name='Calibri',
            size=11,
            color='FF000000',
            bold=False,
            italic=False,
            vertAlign=None,
            underline='none',
            strike=False)
sheet['A1'].font = font
# 填充样式
from openpyxl.styles import PatternFill
# fill_type 的样式为 None 或 solid
cell.fill = PatternFill(fill_type=cell.fill.fill_type, fgColor=cell.fill.fgColor)
from openpyxl.styles import Border, Side
# 边框样式
border = Border(left=Side(border_style=None, color='FF000000'),
                right=Side(border_style=None, color='FF000000'),
                top=Side(border_style=None, color='FF000000'),
                bottom=Side(border_style=None, color='FF000000'),
                diagonal=Side(border_style=None, color='FF000000'),
                diagonal_direction=0,
                outline=Side(border_style=None, color='FF000000'),
                vertical=Side(border_style=None, color='FF000000'),
                horizontal=Side(border_style=None, color='FF000000')
)
# 对齐 horizontal 的值有：distributed, justify, center, left, fill, centerContinuous, right, general
# vertical 的值有：bottom, distributed, justify, center, top
from openpyxl.styles import Alignment
alignment=Alignment(horizontal='general',
                    vertical='bottom',
                    text_rotation=0,
                    wrap_text=False,
                    shrink_to_fit=False,
                    indent=0)
# 整行或整列应用样式
# 合并的单元格可以想象成左上角的单元格来操作。
col = sheet.column_dimensions['A']
col.font = Font(bold=True)
row = sheet.row_dimensions[1]
row.font = Font(underline="single")

wbook.save('../resource/OpenPyxlTest.xlsx')